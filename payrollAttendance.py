import base64
from ast import For
from distutils.log import error
from time import time
import openpyxl as xl
import os
import pandas as pd
from warnings import filterwarnings
from datetime import datetime
import time
from colorama import Fore, Style, Back
import streamlit as st
from payrollWages import wages
import streamlit.components.v1 as components
filterwarnings("ignore")

st.set_page_config("Payroll Attendance", "Picture 1.png" , layout='wide')

with open("style.css") as style:
    st.markdown(f"<style>{style.read()}</style>", unsafe_allow_html=True)

#! ################ THEMING ################################

@st.experimental_memo
def get_img_as_base64(file):
    with open(file, "rb") as f:
        data = f.read()
    return base64.b64encode(data).decode()
img = get_img_as_base64("Picture 1.png")
img_sidebar = get_img_as_base64("lindos.webp")
img_new = get_img_as_base64("nnew.webp")
page_bg_img = f"""
<style>
[data-testid="stAppViewContainer"] > .main {{
background-image: url("data:image/png;base64,{img}");
background-size: 15%;
background-position: 100% 8%;
background-repeat: no-repeat;
background-attachment: fixed;
}}

[data-testid="stSidebar"] > div:first-child {{
background-image: url("data:image/png;base64,{img_new}");
background-size: 80%;
background-position: -180% 70%; 
background-repeat: no-repeat;
background-attachment: fixed;
}}
"""

st.markdown(page_bg_img, unsafe_allow_html=True)

#! ################ PAGES ################################
sb = st.sidebar
sb.header("CHOOSE PROJECT")
project = sb.selectbox("Options", options=["BI Preview","Payroll Project"])
if project == "Payroll Project":
    task = sb.selectbox("Choose Task", options=["Payroll Attendance", "Payroll Wages"])
    if task == "Payroll Attendance":
        @st.cache(show_spinner=False)
        def analysis(files):
            counter = 0
            skipped_files = []
            while True:
                print("\n\nCreating Payroll_Attendance.xlsx\nPlease wait...")
                try:
                    dataframes_to_concat = []
                    start = time.time()
                    for file in files:
                        with st.spinner(f"Processing {counter} out of {len(files)}..."):
                            print(f'Checked {counter} out of {len(files)} files...')
                            # if root != r"C:\Users\lefteris.fthenos\Desktop\attendances\3300" and root != r"C:\Users\lefteris.fthenos\Desktop\attendances\3300\3303" and root != r"C:\Users\lefteris.fthenos\Desktop\attend_test\3300\3302":
                            #    file[-8:-5]
                            try:
                                print(file.name[-11:-5])
                                int(file.name[
                                    -8:-6])  # if the file ends in anything else than a number (i.e. 20220519) and it end in (20220519_v01) the program will skip it

                                # break
                                wb = xl.load_workbook(file)
                                ws = wb["ST21"]
                                hotel_num = ws["A1"].value[:4]
                                date = ws['A3'].value.strftime('%d/%m/%Y')
                                # print(hotel_num)
                                ds = pd.read_excel(file, sheet_name="ST01", engine='openpyxl',
                                                   skiprows=5, converters={'VAT': str})
                                ds.dropna(subset=['VAT'], inplace=True)
                                # ds.replace(pd.NA, "-")

                                ds.drop(axis=1,
                                        labels=['S/N', 'Last Name', 'First Name', 'Starting\n Date', 'Ending\n Date',
                                                'Cost Center'], inplace=True)
                                # print(ds.columns)
                                ds.rename(columns={'Specialty': 'WPDescr', 'Part Time': "PartTime",
                                                   'Correct\nPunch\n(1/0)': 'CorrectPunch', 'Active\n(1/0)': 'Active',
                                                   'Extra\n(1/0)': 'Extra',
                                                   '                    ACTUAL                                       \n': 'Actual',
                                                   'Wage\n(1,0)': 'Wage', 'Overtime\n(Hours)': 'Overtime',
                                                   "Attendance's\nJustification": 'AttendanceJustification',
                                                   'Overtime\nJustification': 'OvertimeJustification',
                                                   'Absence\nJustification': 'AbsenceJustification'}, inplace=True)
                                ds['HotelCode'] = [hotel_num for i in range(0, len(ds.index))]
                                ds['Date'] = [date for j in range(0, len(ds.index))]
                                # print(ds.columns)
                                ds = ds[
                                    ['HotelCode', 'Date', 'WPDescr', 'Sector', 'VAT', 'PartTime', 'CorrectPunch', 'Active',
                                     'Extra',
                                     'Actual', 'Wage', 'Overtime', 'AttendanceJustification',
                                     'OvertimeJustification', 'AbsenceJustification', 'Comments']]
                                dataframes_to_concat.append(ds)
                            except Exception as e:  #
                                skipped_files.append(file.name)
                                print(f"Skipped file {file.name} as it was not valid.")
                                # time.sleep(1)
                                pass
                            counter += 1
                    total_ds = pd.concat(dataframes_to_concat)
                    # print(total_ds)
                    total_ds.to_excel("Payroll_Attendance.xlsx",
                                      sheet_name="Payroll_Attendance", index=False)
                    end = time.time()
                    print('\nThe file was successfully created!')
                    print(f'\nTime spent: {round(end - start, 2)}"')
                    return skipped_files, round(end - start, 2)
                    #break
                except Exception as e:
                    st.warning(f'An error occured: {e}.\nPlease contact Lefteris Fthenos.\n\nPress any key to exit.')
                    exit = input(f'An error occured: {e}.\nPlease contact Lefteris Fthenos.\n\nPress any key to exit.')
                    break

        st.title("Payroll Attendance Files' Processor")
        st.info("You need to delete the file 'Payroll_Attendance.xlsx' from the folder, before you begin the process, otherwise the program will eventually crash.")
        files = st.file_uploader("Upload the Daily Attendance files of the day.\nFiles must be named as 'HLCode - HLName- Daily Attendance_22xxxx.xlsx'  ", type=["xls", "xlsx", "xlsm"], accept_multiple_files=True)
        if files:
            filenames = [file.name for file in files]
            filenames_set = set(filenames)
            if len(filenames) != len(filenames_set):
                st.warning("Warning - You've uploaded duplicate files! Please check them again before running the process.")
            else:
                submit = st.button("Begin the process")
                if submit:
                    skipped_files, time_spent = analysis(files)
                    if skipped_files:
                        st.error(f"Warning - Skipped the following files since they were not valid:")
                        st.warning(f'{skipped_files}"')
                    st.success(f'Time Spent: {round(time_spent,2)}"')
                    st.download_button("Press to download PayrollAttendance.xlsx", data=open("Payroll_Attendance.xlsx", 'rb'), file_name="PayrollAttendance.xlsx")
    elif task == "Payroll Wages":
        st.title("Payroll Wages Files' Processor")
        files = st.file_uploader(
            "Upload the final month's wages xlsx file",
            type=["xls", "xlsx", "xlsm"], accept_multiple_files=True)
        if files:
            filenames = [file.name for file in files]
            filenames_set = set(filenames)
            if len(filenames) != len(filenames_set):
                st.warning(
                    "Warning - You've uploaded duplicate files! Please check them again before running the process.")
            else:
                submit = st.button("Begin the process")
                if submit:
                    time_spent = wages(files)
                    st.success(f'Time Spent: {round(time_spent, 2)}"')
                    st.info("In case you cannot download the file, please run the process again.")
                    st.download_button("Press to download PayrollWages.xlsx",
                                       data=open("PayrollWages.xlsx", 'rb'), file_name="PayrollWages.xlsx")



else:
    components.html(
        """
        <div class="lia-message-custom-report-container">

                <iframe referrerpolicy="no-referrer-when-downgrade" width="1300" height="600" frameborder="0" src="https://app.powerbi.com/view?r=eyJrIjoiZTE0Nzk0MzctZGUxMi00ZmY5LWE2NzItMDUwMjc5ZWYxZGM4IiwidCI6ImUxNzc4MGM2LTczMDgtNDc3YS1iNzY3LWZhNTY5ODI0NWUyYSJ9&amp;embedImagePlaceholder=true" title="iframe"></iframe></div>
        """
        , height=700, width=1300, scrolling=True)

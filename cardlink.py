import datetime
import streamlit as st
import pandas as pd
import openpyxl as xl
from colorama import Fore, Style
import xlsxwriter
from datetime import date
from st_aggrid import AgGrid, JsCode
from st_aggrid.grid_options_builder import GridOptionsBuilder
import warnings
from openpyxl.styles import Color, PatternFill, Font, Border, NamedStyle, Alignment
from openpyxl.styles import colors

@st.cache(show_spinner=True)
def aggrid_customization(df, column_name):
    cellsytle_jscode = JsCode(
        """
    function(params) {
        if (params.value.includes('Eurobank')) {
            return {
                'color': 'white',
                'backgroundColor': 'darkred'
            }
        } else {
            return {
                'color': 'white',
                'backgroundColor': 'blue'
            }
        }
    };
    """
    )
    gb = GridOptionsBuilder.from_dataframe(df)

    # gb.configure_pagination()
    gb.configure_side_bar()
    gb.configure_default_column(value=True, aggFunc="sum",
                                editable=True)
    gb.configure_column(column_name, cellStyle=cellsytle_jscode)

    gridOptions = gb.build()

    AgGrid(df, gridOptions=gridOptions, enable_enterprise_modules=True, theme='streamlit',
           allow_unsafe_jscode=True, fit_columns_on_grid_load=False, key=f"blank{datetime.datetime.now()}")

#@st.cache(show_spinner=True)
def cardlink_auto():
    ekkatharisi_font = Font(size=16, bold=True, italic=True)

    st.title('Cardlink - Apollo')
    st.markdown("""**<u>Οδηγίες Χρήσης**</u>""", unsafe_allow_html=True)
    st.markdown("""
            Με το παρόν πρόγραμμα μπορείτε να διευκολυνθείτε στην άντληση στοιχείων απο τα excel που προκύπτουν απο το [Cardlink - Apollo](https://apollo.cardlink.gr).  \nΑφότου ανεβάσετε το αρχείο excel που εξάγεται απο την ιστοσελίδα **<u>αυτούσιο</u>**, γίνεται μια ανάλυση με αναδιανομή και ταξινόμηση των στηλών του excel για την καλύτερη εξυπηρέτηση σας.
            Στο τέλος του αρχείου excel που θα προκύψει, υπάρχει επιπλέον μια συγκεντρωτική κατάσταση ανά τράπεζα, ανά τύπο κάρτας και ανα πακέτο.
            <u>Για να κατεβάσετε το τροποποιημένο αρχείο κατευθυνθείτε στο τέλος της σελίδας.</u>
            \n**Note:** Στο τέλος της παρούσας σελίδας εμφανίζονται οι συναλλαγές - εάν υφίστανται - που πραγματοποιήθηκαν επιτυχώς και δεν εκκαθαρίστηκαν εγκαίρως.
            """, unsafe_allow_html=True)
    st.info(
        'Tip: Για δική σας εξυπηρέτηση και εξαγωγή ορθών αποτελεσμάτων είναι προτιμότερο να ανεβάζετε αρχεία με τουλάχιστον 3 ημερομηνίες συναλλαγών (π.χ. Χ1/ΧΧ/ΧΧXX, X2/XX/XXXX, X3/XX/XXXX).')
    card_xl = st.file_uploader(
        'Παρακαλώ ανεβάστε αυτούσιο το excel file όπως προκύπτει από την ιστοσελίδα του https://apollo.cardlink.gr',
        type=['xlsx', "xls"], accept_multiple_files=False)
    try:
        if card_xl is not None:
            with st.expander("Expand for file analysis"):
                df_card = pd.read_excel(card_xl)
                # print(df_card)

                # ------- cleaning the data and customizing the df -----------------

                df_card = df_card[
                    ['Ημερομηνία/Date', 'Ώρα/Time', 'Πάροχος Πληρωμών/Payment Provider', 'Αριθμός Κάρτας/Card Number',
                     'Ποσό/Amount', 'Τύπος Κάρτας/Card Type', 'Αποστολή για Εκκαθάριση/In Settlement File',
                     'Απάντηση/Response', 'Αριθμός Πακέτου/Batch Number', "Κλείσιμο Πακέτου/Close Batch",
                     'Κωδικός Επιχείρησης/Merchant ID',
                     'Κωδικός POS/Terminal ID', 'vMerchant ID', 'vTerminal ID', 'Ακύρωση/Voided On',
                     'Ακύρωση από/Voided By', 'Κωδικός Συνδρομητή/Payment Info',
                     'Κωδικός Δραστηριότητας Επιχείρησης/Merchant Category Code',
                     'Συνολικός Αριθμός Δόσεων/Number of Total Instalments', 'Τύπος Συναλλαγής/Transaction Type',
                     'Χαρακτηριστικό Συναλλαγής/Transaction Attribute', 'RRN', 'Ποσό Εξαργύρωσης/Redemption Amount',
                     'Συνολικό Ποσό συναλλαγής/Total Transaction Amount',
                     'Ποσό Φιλοδωρήματος/Tips Amount', 'e-Commerce Order Id', 'Loyalty ID',
                     'Ποσό Πίστωσης (επιχείρηση)/Net credit amount (merchant)', 'Συνολική Προμήθεια/Total commission',
                     'Προμήθεια Cardlink one/Cardlink one commission', 'Προμήθεια Processor/Processor commission',
                     'Card Entry Mode', 'Cardholder Verification Mode', 'Duplicate Transaction Id', 'Εναλλακτική Ονομασία Επιχείρησης/Merchant Alias', 'Εναλλακτική Ονομασία POS/Terminal Alias',
                     'Κωδικός Έγκρισης/Authorization Code', 'Αριθμός Ταμία/Cashier Number',
                     'Αριθμός Ταμείου/Till Number', 'Πρόσθετο πεδίο 1/Info Field 1', 'Πρόσθετο πεδίο 2/Info Field 2',
                     'Πρόσθετο πεδίο 3/Info Field 3',
                     'Ημ/νία Εντολής Πίστωσης προς Τράπεζα Εμπόρου/Payment sent to Merchant’s Bank date',
                     'Ημ/νία Πίστωσης από Τράπεζα Εμπόρου/Merchant’s Bank credit date',
                     'Απάντηση Τράπεζας Εμπόρου/Merchant’s Bank response']]
                df_card.sort_values(
                    by=["Αποστολή για Εκκαθάριση/In Settlement File", 'Πάροχος Πληρωμών/Payment Provider',
                        'Τύπος Κάρτας/Card Type'], inplace=True, ascending=True)
                df_card.set_index('Ημερομηνία/Date', inplace=True)
                df_card.to_excel('Cardlink_customized.xlsx', sheet_name="POS Transactions")
                df_card.reset_index(inplace=True)
                # time.sleep(1)

                # -------------------------------------------------------------------------------
                badge_dict = {}
                found_blanks1 = False
                found_blanks2 = False

                # ! για να βρουμε τις ημερομηνίες εκκαθαρισης ΚΑΙ τα πακετα που εκκαθαρίστηκαν σε κάθε μοναδική ημερομηνια #####################
                print(f'{Fore.BLUE}BEGIN{Style.RESET_ALL}')
                for idx, settlement_date in enumerate(df_card['Αποστολή για Εκκαθάριση/In Settlement File'].unique()):

                    if isinstance(settlement_date,
                                  str):  # an alternative way to drop na values from excel - αν η ημερομηνια εκκαθάρισης ΔΕΝ ΕΙΝΑΙ ΚΕΝΗ
                        badge_list = []
                        for row in df_card.iterrows():
                            if (row[1][6] == settlement_date) and (row[1][8] not in badge_list):
                                badge_list.append(row[1][8])
                        badge_dict[settlement_date] = badge_list
                print(f"{Fore.MAGENTA}Badge dict:{Style.RESET_ALL}{sorted(badge_dict.keys())}") #!ημερομηνιες εκκαθαρισης

                # ! ############################################################################################################################

                if badge_dict:  # αν έχει όντως τιμες - δηλαδη εαν υπαρχουν ημερομηνιες εκκαθαρισης

                    for date in badge_dict.keys():
                        df_settlement_date = pd.DataFrame(
                            columns=['Πακέτο', 'MasterCard', 'Visa', "Maestro", "Amex", "Σύνολο Πακέτου"])

                        print(f'Ημερομηνία εκκαθάρισης συναλλαγών: {date}')
                        for badge in badge_dict[date]:
                            print(f'Έλεγχος πακέτου: {badge}...')
                            master = 0
                            visa = 0
                            maestro = 0
                            amex = 0

                            for row in df_card.iterrows():
                                # print(row[1])
                                if (row[1][8] == badge) and (row[1][5] == "visa") and (
                                        row[1][7] == 'Approved - Εγκρίθηκε') and (row[1][6] == date):
                                    visa += row[1][4]
                                elif (row[1][8] == badge) and (row[1][5] == "mastercard") and (
                                        row[1][7] == 'Approved - Εγκρίθηκε') and (row[1][6] == date):
                                    master += row[1][4]
                                elif (row[1][8] == badge) and (row[1][5] == "maestro") and (
                                        row[1][7] == 'Approved - Εγκρίθηκε') and (row[1][6] == date):
                                    maestro += row[1][4]
                                elif (row[1][8] == badge) and (row[1][5] == "american-express") and (
                                        row[1][7] == 'Approved - Εγκρίθηκε') and (row[1][6] == date):
                                    amex += row[1][4]
                            df_settlement_date = df_settlement_date.append(
                                {'Πακέτο': str(badge), 'MasterCard': master, 'Visa': visa, "Maestro": maestro,
                                 "Amex": amex, "Σύνολο Πακέτου": master + visa + amex + maestro}, ignore_index=True)

                        df_settlement_date.set_index('Πακέτο', inplace=True)
                        workbook = xl.load_workbook('Cardlink_customized.xlsx')
                        ws = workbook.active

                        ws[f'A{ws.max_row + 2}'] = f"Ημερομηνία Εκκαθάρισης: {date}"
                        ekkatharisi = ws[f'A{ws.max_row}']
                        ekkatharisi.font = ekkatharisi_font
                        ekkatharisi.alignment = Alignment(horizontal='center', vertical='center')
                        ekkatharisi.fill = PatternFill(start_color="0099CCFF", end_color="FFC7CE", fill_type="solid")
                        ws.merge_cells(f"A{ws.max_row}:F{ws.max_row}")
                        ws[f'A{ws.max_row + 1}'] = f"Ανάλυση ανά Πακέτο"
                        cell = ws[f'A{ws.max_row}']
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(size=13, bold=True)
                        cell.fill = PatternFill(start_color="00FFFF99", fill_type="solid")
                        ws.merge_cells(f"A{ws.max_row}:F{ws.max_row}")

                        max_row = ws.max_row
                        workbook.save('Cardlink_customized.xlsx')

                        print(df_settlement_date)

                        df_sum = pd.DataFrame(
                            columns=["Σύνολο MasterCard", "Σύνολο Visa", "Σύνολο Maestro", 'Σύνολο Amex',
                                     "Ποσό Εκκαθάρισης"])
                        df_sum = df_sum.append({"Σύνολο MasterCard": df_settlement_date.sum()["MasterCard"],
                                                "Σύνολο Visa": df_settlement_date.sum()["Visa"],
                                                "Σύνολο Maestro": df_settlement_date.sum()["Maestro"],
                                                'Σύνολο Amex': df_settlement_date.sum()["Amex"],
                                                "Ποσό Εκκαθάρισης": df_settlement_date.sum()["Σύνολο Πακέτου"]},
                                               ignore_index=True)
                        df_sum.set_index("Σύνολο MasterCard", inplace=True)
                        print(df_sum)

                        with pd.ExcelWriter('Cardlink_customized.xlsx', mode='a', engine="openpyxl",
                                            if_sheet_exists="overlay") as writer:
                            df_settlement_date.to_excel(writer, startrow=max_row, sheet_name="POS Transactions")

                        workbook = xl.load_workbook('Cardlink_customized.xlsx')
                        ws = workbook.active
                        max_row = ws.max_row
                        ws[f'A{max_row + 1}'] = "Γενικά Σύνολα"
                        ws[f'A{max_row + 1}'].alignment = Alignment(horizontal='center')
                        ws[f'A{max_row + 1}'].font = Font(bold=True)
                        ws[f'B{max_row + 1}'] = df_settlement_date.sum()["MasterCard"]
                        ws[f'B{max_row + 1}'].font = Font(bold=True)
                        ws[f'C{max_row + 1}'] = df_settlement_date.sum()["Visa"]
                        ws[f'C{max_row + 1}'].font = Font(bold=True)
                        ws[f'D{max_row + 1}'] = df_settlement_date.sum()["Maestro"]
                        ws[f'D{max_row + 1}'].font = Font(bold=True)
                        ws[f'E{max_row + 1}'] = df_settlement_date.sum()["Amex"]
                        ws[f'E{max_row + 1}'].font = Font(bold=True)
                        ws[f'F{max_row + 1}'] = df_settlement_date.sum()["Σύνολο Πακέτου"]
                        ws[f'F{max_row + 1}'].font = Font(bold=True, size=14)
                        workbook.save('Cardlink_customized.xlsx')

                        # with pd.ExcelWriter('Cardlink_customized.xlsx', mode='a', engine="openpyxl", if_sheet_exists="overlay") as writer:
                        #    df_sum.to_excel(writer, startrow=max_row+1, startcol=1 , sheet_name="POS Transactions")

                        # ! ################ ΑΝΑΛΥΣΗ ΑΝΑ ΤΡΑΠΕΖΑ `################################################################

                        alpha_amex = 0
                        alpha_maestro = 0
                        alpha_master = 0
                        alpha_visa = 0
                        eurobank_amex = 0
                        eurobank_visa = 0
                        eurobank_master = 0
                        eurobank_maestro = 0
                        bagde_close_df = pd.DataFrame(
                            columns=["Πακέτο", "Πάροχος", "Τύπος Κάρτας", "Συναλλαγή", "Ημ/νία Εκκαθάρισης",
                                     "Κλείσιμο Πακέτου"])
                        # μεταβλητες για συναλλαγές που δεν εκκαθαρίστηκαν σωστα
                        alpha_master_false = 0
                        alpha_maestro_false = 0
                        alpha_visa_false = 0
                        alpha_amex_false = 0
                        eurobank_master_false = 0
                        eurobank_amex_false = 0
                        eurobank_maestro_false = 0
                        eurobank_visa_false = 0

                        badge_list = []

                        for row in df_card.iterrows():
                            # print(row[1])
                            # print(row[1][8])
                            if (row[1][2] == 'NEXI') and (row[1][7] == 'Approved - Εγκρίθηκε') and (
                                    row[1][6] == date):

                                if row[1][5] == 'american-express':
                                    alpha_amex += row[1][4]
                                elif row[1][5] == 'maestro':
                                    alpha_maestro += row[1][4]
                                elif row[1][5] == 'mastercard':
                                    alpha_master += row[1][4]
                                elif row[1][5] == 'visa':
                                    alpha_visa += row[1][4]
                                if row[1][8] not in badge_list:
                                    badge_list.append(row[1][8])

                            elif (row[1][2] == 'Worldline') and (row[1][7] == 'Approved - Εγκρίθηκε') and (
                                    row[1][6] == date): # worline = eurobank
                                # print(row[1][4])
                                if row[1][5] == 'american-express':
                                    eurobank_amex += row[1][4]
                                elif row[1][5] == 'maestro':
                                    eurobank_maestro += row[1][4]
                                elif row[1][5] == 'mastercard':
                                    eurobank_master += row[1][4]
                                elif row[1][5] == 'visa':
                                    eurobank_visa += row[1][4]

                                if row[1][8] not in badge_list:
                                    badge_list.append(row[1][8])

                                if (datetime.datetime.strptime(row[1][9], '%d/%m/%Y %H:%M:%S').time() > datetime.time(0,
                                                                                                                      0,
                                                                                                                      0)) and (
                                        datetime.datetime.strptime(row[1][9],
                                                                   '%d/%m/%Y %H:%M:%S').time() < datetime.time(3, 0,
                                                                                                               0)):

                                    print(f'Το πακέτο {row[1][8]} έκλεισε μετά τις 24:00!')
                                    trans = {"Πακέτο": row[1][8], "Πάροχος": row[1][2], "Τύπος Κάρτας": row[1][5],
                                             "Συναλλαγή": row[1][4], "Ημ/νία Εκκαθάρισης": row[1][6],
                                             "Κλείσιμο Πακέτου": row[1][9]}
                                    if trans["Πάροχος"] == "NEXI" and trans["Τύπος Κάρτας"] == "mastercard":
                                        alpha_master_false += float(trans["Συναλλαγή"])
                                    elif trans["Πάροχος"] == "NEXI" and trans["Τύπος Κάρτας"] == "visa":
                                        alpha_visa_false += float(trans["Συναλλαγή"])
                                    elif trans["Πάροχος"] == "NEXI" and trans["Τύπος Κάρτας"] == "maestro":
                                        alpha_maestro_false += float(trans["Συναλλαγή"])
                                    elif trans["Πάροχος"] == "NEXI" and trans[
                                        "Τύπος Κάρτας"] == "american-express":
                                        alpha_amex_false += float(trans["Συναλλαγή"])
                                    elif trans["Πάροχος"] == "Wordline" and trans["Τύπος Κάρτας"] == "mastercard":
                                        eurobank_master_false += float(trans["Συναλλαγή"])
                                    elif trans["Πάροχος"] == "Wordline" and trans["Τύπος Κάρτας"] == "visa":
                                        eurobank_visa_false += float(trans["Συναλλαγή"])
                                    elif trans["Πάροχος"] == "Wordline" and trans["Τύπος Κάρτας"] == "maestro":
                                        eurobank_maestro_false += float(trans["Συναλλαγή"])
                                    elif trans["Πάροχος"] == "Wordline" and trans["Τύπος Κάρτας"] == "american-express":
                                        eurobank_amex_false += float(trans["Συναλλαγή"])

                                    bagde_close_df = bagde_close_df.append(trans, ignore_index=True)

                            badge_dict[date] = badge_list

                        st.markdown(f"""***<head><font size="-5"><h6>Ημερομηνία εκκαθάρισης συναλλαγών: {date} </h6></font></head>***""", unsafe_allow_html=True)
                        st.markdown(f"""
                        * Αριθμοί Πακέτων: ***{", ".join([str(i) for i in badge_list])}***
                        * Συνολικό ποσό που εκκαθαρίστηκε: ***{(round(alpha_master + alpha_visa + alpha_amex + alpha_maestro + eurobank_visa + eurobank_amex + eurobank_master + eurobank_maestro, 2)).__format__(',.2f')} €*** 
                                                        """)

                        col2, col3 = st.columns((1, 1))

                        workbook = xl.load_workbook('Cardlink_customized.xlsx')
                        ws = workbook.active

                        max_row = ws.max_row
                        ws.merge_cells(f'A{max_row + 1}:F{max_row + 1}')
                        cell = ws[f'A{max_row + 1}']
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = cell.font.copy(size=13, bold=True)
                        cell.fill = PatternFill(start_color="00FFFF99", fill_type="solid")

                        ws[
                            f'A{max_row + 1}'] = f'Ανάλυση ανά τράπεζα - Αριθμοί πακέτων: {", ".join([str(i) for i in badge_list])}'
                        ws[f'A{max_row + 2}'] = 'Γενικό Σύνολο Alpha Bank'
                        ws[f'A{max_row + 3}'] = 'Alpha Bank - AMEX'
                        ws[f'A{max_row + 4}'] = 'Alpha Bank - MASTERCARD'
                        ws[f'A{max_row + 5}'] = 'Alpha Bank - VISA'
                        ws[f'A{max_row + 6}'] = 'Alpha Bank - MAESTRO'

                        # ------------- for streamlit --------------------
                        df_alpha = pd.DataFrame(
                            columns=['Ημερομηνία Εκκαθάρισης', 'Γενικό Σύνολο Alpha Bank', 'Alpha Bank - AMEX',
                                     'Alpha Bank - MASTERCARD', 'Alpha Bank - VISA', 'Alpha Bank - MAESTRO'])
                        df_alpha = df_alpha.append({'Ημερομηνία Εκκαθάρισης': date,
                                                    'Γενικό Σύνολο Alpha Bank': alpha_master + alpha_visa + alpha_amex + alpha_maestro,
                                                    'Alpha Bank - AMEX': alpha_amex,
                                                    'Alpha Bank - MASTERCARD': alpha_master,
                                                    'Alpha Bank - VISA': alpha_visa,
                                                    'Alpha Bank - MAESTRO': alpha_maestro},
                                                   ignore_index=True)
                        # print(df_alpha.transpose())
                        df_alpha.set_index('Ημερομηνία Εκκαθάρισης', inplace=True)
                        df_alpha = df_alpha.transpose().style.format("{:.2f}")
                        col2.dataframe(df_alpha.highlight_max(color='green'))

                        bagde_close_df.sort_values(by=["Πακέτο", "Τύπος Κάρτας"], inplace=True)
                        print(bagde_close_df)
                        badges_array = bagde_close_df['Πακέτο'].unique()
                        badges_list = badges_array.tolist()

                        if badges_list:
                            if datetime.datetime.strptime(date, '%d/%m/%Y').weekday() == 0:
                                st.warning(f""" 
                                ***Προσοχή!!!*** Το/τα πακέτο/τα {badges_list} έκλεισε/αν μετά τις 00:00. Οι παρακάτω συναλλαγές ***εκκαθαρίστηκαν διαφορετική ημερομηνία από την αναμενόμενη***.
                                Δώστε ιδιαίτερη προσοχή στο γεγονός ότι η ημερομηνία εκκαθάρισης είναι Δευτέρα, συνεπώς εαν οι συναλλαγές του πακέτου που δεν έκλεισε σωστά είναι Σαββάτου θα φανούν στην τράπεζα την Δευτέρα, ειδάλλως αν είναι Κυριακής, θα φανούν την Τρίτη.""")
                            else:
                                st.warning(
                                    f"""***Προσοχή!!!*** Το/τα πακέτο/τα {badges_list} έκλεισε/αν μετά τις 00:00. Οι παρακάτω συναλλαγές ***εκκαθαρίστηκαν διαφορετική ημερομηνία από την αναμενόμενη***.""")

                            st.warning(f"""
                                                ***Σύνολο Eurobank :*** {eurobank_maestro_false +
                                                                         eurobank_amex_false + eurobank_visa_false + eurobank_master_false}
                                                * *Eurobank Mastercard:* {eurobank_master_false}
                                                * *Eurobank Visa:* {eurobank_visa_false}
                                                * *Eurobank Maestro:* {eurobank_maestro_false}
                                                * *Eurobank Amex:* {eurobank_amex_false}
                                                ***Σύνολο Alpha Bank :*** {alpha_amex_false + alpha_master_false + alpha_maestro_false + alpha_visa_false}
                                                * *Alpha Mastercard:* {alpha_master_false}
                                                * *Alpha Visa:* {alpha_visa_false}
                                                * *Alpha Maestro:* {alpha_maestro_false}
                                                * *Alpha Amex:* {alpha_amex_false}
                                                            """)
                            aggrid_customization(bagde_close_df, 'Πάροχος')

                        # ------------------------------------------------------

                        ws[f'B{max_row + 2}'] = alpha_master + alpha_visa + alpha_amex + alpha_maestro
                        alpha_total = ws[f'B{max_row + 3}']
                        alpha_total.font = Font(bold=True)
                        # alpha_total.fill = Fill()
                        ws[f'B{max_row + 3}'] = alpha_amex
                        ws[f'B{max_row + 4}'] = alpha_master
                        ws[f'B{max_row + 5}'] = alpha_visa
                        ws[f'B{max_row + 6}'] = alpha_maestro

                        ws[f'D{max_row + 2}'] = 'Γενικό Σύνολο Eurobank'
                        ws[f'D{max_row + 3}'] = 'Eurobank - AMEX'
                        ws[f'D{max_row + 4}'] = 'Eurobank - MASTERCARD'
                        ws[f'D{max_row + 5}'] = 'Eurobank - VISA'
                        ws[f'D{max_row + 6}'] = 'Eurobank - MAESTRO'

                        ws[f'E{max_row + 2}'] = eurobank_visa + eurobank_amex + eurobank_master + eurobank_maestro

                        eurobank_total = ws[f'E{max_row + 2}']
                        eurobank_total.font = Font(bold=True)
                        ws[f'E{max_row + 3}'] = eurobank_amex
                        ws[f'E{max_row + 4}'] = eurobank_master
                        ws[f'E{max_row + 5}'] = eurobank_visa
                        ws[f'E{max_row + 6}'] = eurobank_maestro

                        # ------------- for streamlit --------------------
                        df_euro = pd.DataFrame(
                            columns=['Γενικό Σύνολο Eurobank', 'Eurobank - AMEX', 'Eurobank - MASTERCARD',
                                     'Eurobank - VISA', 'Eurobank - MAESTRO'])
                        df_euro = df_euro.append({'Ημερομηνία Εκκαθάρισης': date,
                                                  'Γενικό Σύνολο Eurobank': eurobank_visa + eurobank_amex + eurobank_master + eurobank_maestro,
                                                  'Eurobank - AMEX': eurobank_amex,
                                                  'Eurobank - MASTERCARD': eurobank_master,
                                                  'Eurobank - VISA': eurobank_visa,
                                                  'Eurobank - MAESTRO': eurobank_maestro},
                                                 ignore_index=True)

                        # print(df_euro.transpose())

                        df_euro.set_index('Ημερομηνία Εκκαθάρισης', inplace=True)
                        df_euro = df_euro.transpose().style.format("{:,.2f}")

                        col3.dataframe(df_euro.highlight_max(color='green'))

                        # -------------------------------------------------------
                        workbook.save('Cardlink_customized.xlsx')



                    else:  # αν η εντοπίζονται ημερομηνίες εκκαθάρισης που είναι blank κελιά
                        print('entered here!')
                        sum_incomplete = 0
                        sum_transcations_incomplete1 = 0
                        sum_transcations_incomplete2 = 0
                        sum_alpha1 = 0
                        sum_eurobank1 = 0
                        sum_alpha2 = 0
                        sum_eurobank2 = 0

                        df_incomplete1 = pd.DataFrame(
                            columns=['Ημερομηνία Συναλλαγής', 'Ώρα', "Αριθμός Πακέτου", "Πάροχος Πληρωμών",
                                     "Τύπος Κάρτας",
                                     "Ποσό", "Κλείσιμο Πακέτου", "Έπρεπε να εκκαθαριστεί"])
                        df_incomplete2 = pd.DataFrame(
                            columns=['Ημερομηνία Συναλλαγής', 'Ώρα', "Αριθμός Πακέτου", "Πάροχος Πληρωμών",
                                     "Τύπος Κάρτας",
                                     "Ποσό", "Κλείσιμο Πακέτου", "Έπρεπε να εκκαθαριστεί"])

                        print(badge_dict)
                        for row in df_card.iterrows():
                            # print(row[1][7], ' - ', row[1][6])
                            if (row[1][7] == 'Approved - Εγκρίθηκε') and (not isinstance(row[1][6], str)):
                                # print(row)
                                # print('entered here 2!!!', row[1][9])
                                try:
                                    date_time_close_badge = datetime.datetime.strptime(row[1][9], '%d/%m/%Y %H:%M:%S')
                                    print(date_time_close_badge.time().hour)

                                    if (date_time_close_badge.date().strftime("%d/%m/20%y") in badge_dict.keys()) and (
                                            date_time_close_badge.time() >= datetime.time(0, 0, 0)) and (
                                            date_time_close_badge.time() <= datetime.time(3, 0, 0)):
                                        found_blanks1 = True
                                        sum_incomplete += 1
                                        sum_transcations_incomplete1 += row[1][4]
                                        if row[1][2] == 'Alpha Bank':
                                            sum_alpha1 += row[1][4]
                                        else:
                                            sum_eurobank1 += row[1][4]

                                        print(
                                            f'{Fore.MAGENTA}Found {sum_incomplete} incomplete transactions!{Style.RESET_ALL} - date {date_time_close_badge}')
                                        # print(row)
                                        transaction = {'Ημερομηνία Συναλλαγής': row[1][0], "Ώρα": row[1][1],
                                                       "Αριθμός Πακέτου": row[1][8], "Πάροχος Πληρωμών": row[1][2],
                                                       "Τύπος Κάρτας": row[1][5], "Ποσό": row[1][4],
                                                       "Κλείσιμο Πακέτου": row[1][9],
                                                       "Έπρεπε να εκκαθαριστεί": date_time_close_badge.date().strftime(
                                                           "%d/%m/20%y")}
                                        df_incomplete1 = df_incomplete1.append(transaction, ignore_index=True)

                                    else:
                                        found_blanks2 = True
                                        sum_incomplete += 1
                                        sum_transcations_incomplete2 += row[1][4]
                                        if row[1][2] == 'Alpha Bank':
                                            sum_alpha2 += row[1][4]
                                        else:
                                            sum_eurobank2 += row[1][4]
                                        transaction = {'Ημερομηνία Συναλλαγής': row[1][0], "Ώρα": row[1][1],
                                                       "Αριθμός Πακέτου": row[1][8], "Πάροχος Πληρωμών": row[1][2],
                                                       "Τύπος Κάρτας": row[1][5], "Ποσό": row[1][4],
                                                       "Κλείσιμο Πακέτου": row[1][9],
                                                       "Έπρεπε να εκκαθαριστεί": date_time_close_badge.date().strftime(
                                                           "%d/%m/20%y")}
                                        df_incomplete2 = df_incomplete2.append(transaction, ignore_index=True)


                                except ValueError:

                                    # value error προκύπτει εαν στην στήλη "Κλείσιμο Πακέτου" εξετάζουμε κελί που έχει τιμή "-". Η συγκεκριμένη τιμή (απο τη στιγμή που η συναλλαγή ειναι approved) σημαίνει ότι το πακέτο δεν έχει κλείσει ακόμη και είναι συναλλαγή που θα εκκαθαριστεί την επόμενη μέρα (νέο πακέτο). Δηλαδή η συναλλαγή είναι μια χαρά.

                                    pass
                if found_blanks1:
                    st.subheader('Επιτυχείς συναλλαγές που δεν εκκαθαρίστηκαν λόγω καθυστέρησης κλεισίματος του POS')
                    st.markdown(
                        f'Συνολικό ύψος συναλλαγών που δεν εκκαθαρίστηκαν: ***{sum_transcations_incomplete1} €***')
                    st.markdown(f"""
                                * Alpha Bank: ***{sum_alpha1} €***
                                * Eurobank: ***{sum_eurobank1} €***""")

                    aggrid_customization(df_incomplete1, "Πάροχος Πληρωμών")

                if found_blanks1:
                    my_red = colors.Color(rgb='00FF0000')
                    my_fill = PatternFill(patternType='solid', fgColor='00993366')

                    workbook = xl.load_workbook('Cardlink_customized.xlsx')
                    ws = workbook.active

                    max_row = ws.max_row
                    ws.merge_cells(f'A{max_row + 2}:F{max_row + 2}')

                    cell = ws[f'A{max_row + 2}']
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = cell.font.copy(bold=True, color="00FFFFFF")
                    cell.fill = my_fill
                    ws[
                        f'A{max_row + 2}'] = 'Συναλλαγές που πραγματοποιήθηκαν επιτυχώς και δεν εκκαθαρίστηκαν λόγω καθυστέρησης κλεισίματος πακέτου'
                    ws[f'A{max_row + 3}'] = 'Alpha Bank'
                    ws[f'B{max_row + 3}'] = sum_alpha1
                    ws[f'A{max_row + 4}'] = 'Eurobank'
                    ws[f'B{max_row + 4}'] = sum_eurobank1
                    ws[f'A{max_row + 5}'] = 'Σύνολο'
                    ws[f'B{max_row + 5}'] = sum_eurobank1 + sum_alpha1
                    cell_total = ws[f'B{max_row + 5}']
                    cell_total.font = cell_total.font.copy(bold=True)

                    workbook.save('Cardlink_customized.xlsx')

                if found_blanks2:
                    cellsytle_jscode = JsCode(
                        """
                    function(params) {
                        if (params.value.includes('Eurobank')) {
                            return {
                                'color': 'white',
                                'backgroundColor': 'darkred'
                            }
                        } else {
                            return {
                                'color': 'white',
                                'backgroundColor': 'blue'
                            }
                        }
                    };
                    """
                    )

                    st.subheader('Επιτυχείς συναλλαγές που δεν εκκαθαρίστηκαν για άγνωστη αιτία')
                    st.markdown(
                        f'Συνολικό ύψος συναλλαγών που δεν εκκαθαρίστηκαν: ***{sum_transcations_incomplete2} €***')
                    st.markdown(f"""
                                * Alpha Bank: ***{sum_alpha2} €***
                                * Eurobank: ***{sum_eurobank2} €***""")

                    gb = GridOptionsBuilder.from_dataframe(df_incomplete2)

                    # gb.configure_pagination()
                    gb.configure_side_bar()
                    gb.configure_default_column(value=True, aggFunc="sum",
                                                editable=True)
                    gb.configure_column("Πάροχος Πληρωμών", cellStyle=cellsytle_jscode)

                    gridOptions = gb.build()

                    AgGrid(df_incomplete2, gridOptions=gridOptions, enable_enterprise_modules=True, theme='streamlit',
                           allow_unsafe_jscode=True, fit_columns_on_grid_load=False, key="blank2")

                if found_blanks2:
                    my_red = colors.Color(rgb='00FF0000')
                    my_fill = PatternFill(patternType='solid', fgColor=my_red)

                    workbook = xl.load_workbook('Cardlink_customized.xlsx')
                    ws = workbook.active

                    max_row = ws.max_row
                    ws.merge_cells(f'A{max_row + 2}:F{max_row + 2}')
                    cell = ws[f'A{max_row + 2}']
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = cell.font.copy(bold=True, color="00FFFFFF")
                    cell.fill = my_fill
                    ws[
                        f'A{max_row + 2}'] = 'Συναλλαγές που πραγματοποιήθηκαν επιτυχώς και δεν εκκαθαρίστηκαν για άγνωστη αιτία'
                    ws[f'A{max_row + 3}'] = 'Alpha Bank'
                    ws[f'B{max_row + 3}'] = sum_alpha2
                    ws[f'A{max_row + 4}'] = 'Eurobank'
                    ws[f'B{max_row + 4}'] = sum_eurobank2
                    ws[f'A{max_row + 5}'] = 'Σύνολο'
                    ws[f'B{max_row + 5}'] = sum_eurobank2 + sum_alpha2
                    cell_total = ws[f'B{max_row + 5}']
                    cell_total.font = cell_total.font.copy(bold=True)

                    workbook.save('Cardlink_customized.xlsx')

                #wb = xl.load_workbook('Cardlink_customized.xlsx')
                #ws = wb.active
                # ws.delete_cols(1, 1)
                #wb.save('Cardlink_customized.xlsx')
            st.markdown(''' Κατεβάστε το τροποποιημένο αρχείο excel.''')
            st.download_button('Download XLSX file', data=open('Cardlink_customized.xlsx', 'rb'),
                               file_name=f'Cardlink {datetime.date.today().strftime("%d-%m-%Y")}.xlsx')
    except Exception as e:
        print(e)
        st.error(e)

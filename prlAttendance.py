import pandas as pd
import fnmatch
import streamlit as st
from colorama import Fore, Style
from warnings import filterwarnings
import time
import datetime as dt
import openpyxl as xl
filterwarnings("ignore")
pattern = "* - * - Daily Attendance_2?????.xlsx"


@st.cache(show_spinner=False, suppress_st_warning=True)
def analysis(files):
    wrong_filenames = []
    for file in files:
        #print(fnmatch.fnmatch(file.name, pattern))
        if not fnmatch.fnmatch(file.name, pattern):
            wrong_filenames.append(file.name)
    print(f"{Fore.MAGENTA}{wrong_filenames}{Style.RESET_ALL}")
    if wrong_filenames:
        return wrong_filenames,0
    else:
        counter = 0
        skipped_files = []

        while True:
            print("\n\nCreating Payroll_Attendance.xlsx\nPlease wait...")
            try:
                dataframes_to_concat = []
                start = time.time()
                for file in files:
                    with st.spinner(f"Processing {counter} out of {len(files)}..."):
                        print(f'Checked {counter} out of {len(files)} files...')
                        # if root != r"C:\Users\lefteris.fthenos\Desktop\attendances\3300" and root != r"C:\Users\lefteris.fthenos\Desktop\attendances\3300\3303" and root != r"C:\Users\lefteris.fthenos\Desktop\attend_test\3300\3302":
                        #    file[-8:-5]
                        try:
                            print(file.name[-11:-5])
                            int(file.name[
                                -8:-6])  # if the file ends in anything else than a number (i.e. 20220519) and it end in (20220519_v01) the program will skip it

                            # break
                            wb = xl.load_workbook(file)
                            ws = wb["ST21"]
                            hotel_num = ws["A1"].value[:4]
                            date = ws['A3'].value.strftime('%d/%m/%Y')
                            # print(hotel_num)
                            ds = pd.read_excel(file, sheet_name="ST01", engine='openpyxl',
                                               skiprows=5, converters={'VAT': str})
                            ds.dropna(subset=['VAT'], inplace=True)
                            # ds.replace(pd.NA, "-")

                            ds.drop(axis=1,
                                    labels=['S/N', 'Last Name', 'First Name', 'Starting\n Date', 'Ending\n Date',
                                            'Cost Center'], inplace=True)
                            # print(ds.columns)
                            ds.rename(columns={'Specialty': 'WPDescr', 'Part Time': "PartTime",
                                               'Correct\nPunch\n(1/0)': 'CorrectPunch', 'Active\n(1/0)': 'Active',
                                               'Extra\n(1/0)': 'Extra',
                                               '                    ACTUAL                                       \n': 'Actual',
                                               'Wage\n(1,0)': 'Wage', 'Overtime\n(Hours)': 'Overtime',
                                               "Attendance's\nJustification": 'AttendanceJustification',
                                               'Overtime\nJustification': 'OvertimeJustification',
                                               'Absence\nJustification': 'AbsenceJustification'}, inplace=True)
                            ds['HotelCode'] = [hotel_num for i in range(0, len(ds.index))]
                            ds['Date'] = [date for j in range(0, len(ds.index))]
                            # print(ds.columns)
                            ds = ds[
                                ['HotelCode', 'Date', 'WPDescr', 'Sector', 'VAT', 'PartTime', 'CorrectPunch', 'Active',
                                 'Extra',
                                 'Actual', 'Wage', 'Overtime', 'AttendanceJustification',
                                 'OvertimeJustification', 'AbsenceJustification', 'Comments']]
                            dataframes_to_concat.append(ds)
                        except Exception as e:  #
                            skipped_files.append(file.name)
                            print(f"Skipped file {file.name} as it was not valid.")
                            # time.sleep(1)
                            pass
                        counter += 1
                if dataframes_to_concat:
                    total_ds = pd.concat(dataframes_to_concat)
                    # print(total_ds)
                    total_ds['Overtime'] = total_ds['Overtime'].astype(str)
                    total_ds.to_excel("Payroll_Attendance.xlsx",
                                      sheet_name="Payroll_Attendance", index=False)
                    print('\nThe file was successfully created!')
                end = time.time()
                print(f'\nTime spent: {round(end - start, 2)}"')
                return skipped_files, round(end - start, 2)
                #break
            except Exception as e:
                st.warning(f'An error occured: {e}.\nPlease contact Lefteris Fthenos.\n\nPress any key to exit.')
                #exit = input(f'An error occured: {e}.\nPlease contact Lefteris Fthenos.\n\nPress any key to exit.')
                return [], 0
